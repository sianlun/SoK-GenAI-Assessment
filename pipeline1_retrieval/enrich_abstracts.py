#!/usr/bin/env python3
"""
enrich_abstracts.py — Fetch full abstracts from open APIs, then re-run screening
══════════════════════════════════════════════════════════════════════════════════

PROBLEM
───────
BibTeX exports from academic databases commonly truncate or abbreviate abstracts:
  • Web of Science  — caps at ~250 words; reformats paragraphs
  • Scopus          — silently truncates at ~1,500 characters for older records
  • IEEE            — generally complete, but some conference records are short
  • ACM             — generally complete
  • ERIC            — carries author-submitted draft, may differ from published

Truncated abstracts cause FALSE EXCLUSIONS in keyword screening — if assessment-
type or education-context keywords only appear in the second half of an abstract,
those papers get incorrectly flagged EC-ASSESS or EC-EDU.

THREE-SOURCE ENRICHMENT PIPELINE
─────────────────────────────────
For each record that has a DOI, we query three free, open APIs:

  1. Semantic Scholar Batch API  (primary — best CS/engineering coverage)
     https://api.semanticscholar.org/graph/v1/paper/batch
     Accepts up to 500 DOIs per POST → 2,162 records ≈ 5 calls (very fast).
     Returns clean plain text. No API key required.

  2. CrossRef API  (fallback — broad coverage, JATS XML cleaned automatically)
     https://api.crossref.org/works/{doi}
     Uses 20-thread pool so ~500 misses complete in <15 s.

  3. OpenAlex batch filter API  (final fallback — inverted-index reconstruction)
     https://api.openalex.org/works?filter=doi:A|doi:B&per_page=200
     Batches 50 DOIs per request; abstracts stored as inverted index → rebuilt.

SELECTION POLICY
────────────────
For each record we keep whichever abstract is LONGER: enriched API version or
original BIB export. This ensures we never downgrade an already-complete abstract.

OUTPUT
──────
  _abstract_cache.json           Cache of all API responses (enables --resume)
  r2_Master_List_enriched.xlsx   Master list with enriched abstracts
  enrichment_report.txt          Per-source stats

USAGE
─────
  python3 enrich_abstracts.py
  python3 enrich_abstracts.py --email you@uni.edu   # polite CrossRef pool
  python3 enrich_abstracts.py --resume              # skip cached DOIs
  python3 enrich_abstracts.py --limit 100           # test on first 100
══════════════════════════════════════════════════════════════════════════════════
"""

import argparse
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE          = Path(__file__).parent
MASTER_XLSX   = HERE / "r2_Master_List.xlsx"
ENRICHED_XLSX = HERE / "r2_Master_List_enriched.xlsx"
CACHE_FILE    = HERE / "_abstract_cache.json"
REPORT_FILE   = HERE / "enrichment_report.txt"

S2_BATCH_URL  = "https://api.semanticscholar.org/graph/v1/paper/batch"
CROSSREF_URL  = "https://api.crossref.org/works/{doi}"
OPENALEX_URL  = "https://api.openalex.org/works"
S2_BATCH_SIZE = 500    # S2 allows 500 per POST


# ═══════════════════════════════════════════════════════════════════════════════
# TEXT CLEANING
# ═══════════════════════════════════════════════════════════════════════════════

def clean_jats(text: str) -> str:
    """Strip JATS/HTML XML tags from CrossRef abstracts and decode entities."""
    if not text:
        return ""
    text = re.sub(r"</?jats:[^>]*>", " ", text)
    text = re.sub(r"<[^>]+>", " ", text)
    for ent, ch in [("&amp;","&"),("&lt;","<"),("&gt;",">"),
                    ("&quot;",'"'),("&apos;","'"),("&#39;","'"),("&nbsp;"," ")]:
        text = text.replace(ent, ch)
    return re.sub(r"\s+", " ", text).strip()


def reconstruct_openalex(inverted_index: dict) -> str:
    """Reconstruct an OpenAlex abstract from its inverted index."""
    if not inverted_index:
        return ""
    try:
        max_pos = max(p for positions in inverted_index.values() for p in positions)
        words = [""] * (max_pos + 1)
        for word, positions in inverted_index.items():
            for pos in positions:
                words[pos] = word
        return " ".join(w for w in words if w).strip()
    except Exception:
        return ""


def is_better(new_text: str, old_text: str, min_gain: int = 100) -> bool:
    """Return True if new_text is substantially longer than old_text."""
    return len(new_text.strip()) > len(old_text.strip()) + min_gain


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — Semantic Scholar Batch (fast: ~5 POST calls for 2,162 records)
# ═══════════════════════════════════════════════════════════════════════════════

def phase1_s2(dois: list[str]) -> dict[str, str]:
    """Returns {doi_lower: abstract_text} for all DOIs found in S2."""
    session = requests.Session()
    session.headers["User-Agent"] = "SoK-Pipeline/1.0"
    out: dict[str, str] = {}
    batches = [dois[i:i+S2_BATCH_SIZE] for i in range(0, len(dois), S2_BATCH_SIZE)]
    print(f"    S2: {len(dois)} DOIs → {len(batches)} batches")

    for n, batch in enumerate(batches, 1):
        ids = [f"DOI:{d}" for d in batch]
        try:
            resp = session.post(
                S2_BATCH_URL,
                params={"fields": "abstract"},
                json={"ids": ids},
                timeout=30,
            )
            if resp.status_code == 200:
                for item in resp.json():
                    if not item:
                        continue
                    ext = (item.get("externalIds") or {})
                    doi_key = ext.get("DOI", "").lower()
                    abstract = (item.get("abstract") or "").strip()
                    if doi_key and abstract:
                        out[doi_key] = abstract
            else:
                print(f"      batch {n}: HTTP {resp.status_code}")
        except Exception as e:
            print(f"      batch {n}: error {e}")
        print(f"      batch {n}/{len(batches)} done ({len(out)} found so far)")
        time.sleep(0.5)

    return out


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — CrossRef (threaded, ~20 concurrent workers)
# ═══════════════════════════════════════════════════════════════════════════════

def _crossref_one(doi: str, email: str) -> tuple[str, str]:
    """Fetch one DOI from CrossRef. Returns (doi, abstract)."""
    url = CROSSREF_URL.format(doi=doi)
    params = {"mailto": email} if email else {}
    try:
        resp = requests.get(url, params=params, timeout=12,
                            headers={"User-Agent": "SoK-Pipeline/1.0"})
        if resp.status_code == 200:
            abstract = resp.json().get("message", {}).get("abstract", "") or ""
            return doi, clean_jats(abstract)
    except Exception:
        pass
    return doi, ""


def phase2_crossref(dois: list[str], email: str, workers: int = 20) -> dict[str, str]:
    """Fetch abstracts for a list of DOIs from CrossRef using a thread pool."""
    out: dict[str, str] = {}
    print(f"    CrossRef: {len(dois)} DOIs → {workers} threads")
    found = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(_crossref_one, doi, email): doi for doi in dois}
        done = 0
        for future in as_completed(futures):
            doi, abstract = future.result()
            done += 1
            if abstract and len(abstract) > 80:
                out[doi] = abstract
                found += 1
            if done % 100 == 0:
                print(f"      {done}/{len(dois)} done ({found} found)", flush=True)
    print(f"    CrossRef found: {found}")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — OpenAlex batch filter (50 DOIs per request)
# ═══════════════════════════════════════════════════════════════════════════════

def phase3_openalex(dois: list[str], batch_size: int = 50) -> dict[str, str]:
    """
    Use OpenAlex filter API to batch-fetch abstracts (inverted index).
    Each request handles up to 50 DOIs via pipe-separated filter.
    """
    session = requests.Session()
    session.headers["User-Agent"] = "SoK-Pipeline/1.0"
    out: dict[str, str] = {}
    batches = [dois[i:i+batch_size] for i in range(0, len(dois), batch_size)]
    print(f"    OpenAlex: {len(dois)} DOIs → {len(batches)} batch requests")
    found = 0

    for n, batch in enumerate(batches, 1):
        # OpenAlex filter: doi:X|doi:Y|...
        doi_filter = "|".join(f"doi:{d}" for d in batch)
        try:
            resp = session.get(
                OPENALEX_URL,
                params={
                    "filter": doi_filter,
                    "select": "doi,abstract_inverted_index",
                    "per_page": batch_size,
                },
                timeout=20,
            )
            if resp.status_code == 200:
                for item in (resp.json().get("results") or []):
                    doi_raw = (item.get("doi") or "").lower()
                    # OpenAlex returns full URL: https://doi.org/10.xxx
                    doi_key = re.sub(r"^https?://doi\.org/", "", doi_raw)
                    inv = item.get("abstract_inverted_index") or {}
                    abstract = reconstruct_openalex(inv)
                    if doi_key and abstract:
                        out[doi_key] = abstract
                        found += 1
            else:
                print(f"      batch {n}: HTTP {resp.status_code}")
        except Exception as e:
            print(f"      batch {n}: error {e}")
        time.sleep(0.15)

    print(f"    OpenAlex found: {found}")
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# APPLY ENRICHMENT
# ═══════════════════════════════════════════════════════════════════════════════

def apply_enrichment(records: list[dict], cache: dict) -> list[dict]:
    """Replace BIB abstracts with API versions where the API version is longer."""
    enriched = []
    counts = {"upgraded": 0, "kept_bib": 0, "no_doi": 0, "no_api": 0}

    for rec in records:
        doi = str(rec.get("DOI", "") or "").strip().lower()
        # Normalise DOI: strip URL prefix
        doi = re.sub(r"^https?://doi\.org/", "", doi)

        orig = str(rec.get("Abstract", "") or "").strip()

        if not doi:
            counts["no_doi"] += 1
            enriched.append({**rec, "Abstract_Source": "BIB (no DOI)"})
            continue

        cached = cache.get(doi)
        if not cached:
            counts["no_api"] += 1
            enriched.append({**rec, "Abstract_Source": "BIB (API miss)"})
            continue

        api_text   = cached.get("abstract", "").strip()
        api_source = cached.get("source", "Unknown")

        if is_better(api_text, orig):
            new_rec = dict(rec)
            new_rec["Abstract"]        = api_text
            new_rec["Abstract_Source"] = api_source
            enriched.append(new_rec)
            counts["upgraded"] += 1
        else:
            enriched.append({**rec, "Abstract_Source": "BIB (already complete)"})
            counts["kept_bib"] += 1

    print(f"    Upgraded  : {counts['upgraded']}")
    print(f"    Kept BIB  : {counts['kept_bib']}")
    print(f"    No DOI    : {counts['no_doi']}")
    print(f"    API miss  : {counts['no_api']}")
    return enriched


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

def save_enriched_xlsx(records: list[dict], out_path: Path, orig_path: Path):
    thin   = Side(style="thin", color="B8B8B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR    = PatternFill("solid", fgColor="1F3864")
    GREEN  = PatternFill("solid", fgColor="E2EFDA")
    BLUE   = PatternFill("solid", fgColor="DDEEFF")   # enriched rows

    base_cols = [
        "Record_ID","Record_Status","DB_Source","All_DB_Sources","Dup_Group",
        "Entry_Type","Title","Authors","Year","Source_Title","Publisher",
        "Volume","Issue","Pages","DOI","ISSN","ISBN","Keywords","Abstract",
        "URL","Screening_Status","Exclusion_Reason","Study_Design","AI_Tool",
        "Assessment_Type","Education_Context","Sample_Size","Key_Findings",
        "SoK_Theme","Abstract_Source",
    ]
    all_keys = list(records[0].keys()) if records else []
    cols = [c for c in base_cols if c in all_keys]
    cols += [k for k in all_keys if k not in cols]

    COL_W = {
        "Record_ID":8,"Record_Status":12,"DB_Source":10,"All_DB_Sources":22,
        "Dup_Group":9,"Entry_Type":12,"Title":55,"Authors":35,"Year":6,
        "Source_Title":40,"Publisher":22,"Volume":7,"Issue":7,"Pages":10,
        "DOI":28,"ISSN":12,"ISBN":14,"Keywords":35,"Abstract":60,"URL":30,
        "Screening_Status":14,"Exclusion_Reason":18,"Study_Design":16,
        "AI_Tool":16,"Assessment_Type":18,"Education_Context":18,
        "Sample_Size":10,"Key_Findings":40,"SoK_Theme":20,"Abstract_Source":20,
    }

    # Build workbook — copy non-Unique-Records sheets from original
    wb_orig = load_workbook(orig_path)
    wb_new  = Workbook()

    ws = wb_new.active
    ws.title = "Unique Records"

    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = HDR
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c_idx)].width = COL_W.get(col, 16)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    for r_idx, rec in enumerate(records, 2):
        src  = str(rec.get("Abstract_Source", ""))
        fill = BLUE if src in ("SemanticScholar","CrossRef","OpenAlex") else GREEN
        for c_idx, col in enumerate(cols, 1):
            val = rec.get(col, "")
            if val is None or (isinstance(val, float) and str(val) == "nan"):
                val = ""
            c = ws.cell(row=r_idx, column=c_idx, value=str(val) if val != "" else "")
            c.fill = fill; c.border = border
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top")
        ws.row_dimensions[r_idx].height = 18
    ws.auto_filter.ref = ws.dimensions

    # Copy other sheets
    for sname in wb_orig.sheetnames:
        if sname == "Unique Records":
            continue
        ws_o = wb_orig[sname]
        ws_n = wb_new.create_sheet(title=sname)
        for row in ws_o.iter_rows():
            for cell in row:
                nc = ws_n.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    nc.font = cell.font.copy()
                    nc.fill = cell.fill.copy()
                    nc.border = cell.border.copy()
                    nc.alignment = cell.alignment.copy()
        for cd in ws_o.column_dimensions.values():
            ws_n.column_dimensions[cd.index].width = cd.width

    wb_new.save(out_path)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default=str(MASTER_XLSX))
    parser.add_argument("--out",    default=str(ENRICHED_XLSX))
    parser.add_argument("--email",  default="kiawin@gmail.com",
                        help="Email for CrossRef polite pool")
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--limit",  type=int, default=0)
    args = parser.parse_args()

    print(f"\n{'═'*70}")
    print("  Abstract Enrichment — enrich_abstracts.py")
    print(f"{'═'*70}")
    print(f"  Sources: Semantic Scholar → CrossRef (threaded) → OpenAlex (batch)")

    # Load records
    df = pd.read_excel(args.input, sheet_name="Unique Records", dtype=str)
    df = df.fillna("")
    records = df.to_dict("records")
    if args.limit:
        records = records[:args.limit]

    # Normalise all DOIs
    def norm_doi(s):
        s = str(s or "").strip().lower()
        return re.sub(r"^https?://doi\.org/", "", s)

    all_dois = list({norm_doi(r.get("DOI","")) for r in records
                     if norm_doi(r.get("DOI",""))})
    print(f"  Records : {len(records)}")
    print(f"  DOIs    : {len(all_dois)} unique")

    # Load cache
    cache: dict = {}
    if args.resume and CACHE_FILE.exists():
        cache = json.loads(CACHE_FILE.read_text())
        print(f"  Cache   : {len(cache)} DOIs already resolved")

    cached_dois = set(cache.keys())
    to_fetch    = [d for d in all_dois if d not in cached_dois]
    print(f"  To fetch: {len(to_fetch)} DOIs\n")

    t0 = time.time()

    # ── Phase 1: Semantic Scholar batch ───────────────────────────────────────
    print("  [Phase 1] Semantic Scholar batch")
    if to_fetch:
        s2 = phase1_s2(to_fetch)
        for doi, abstract in s2.items():
            cache[doi] = {"abstract": abstract, "source": "SemanticScholar"}
        print(f"  S2 total: {len(s2)} abstracts fetched  ({time.time()-t0:.1f}s)\n")
    else:
        print("  All DOIs cached — skipping S2\n")
        s2 = {}

    # ── Phase 2: CrossRef for S2 misses ───────────────────────────────────────
    cr_needed = [d for d in to_fetch if d not in cache]
    print(f"  [Phase 2] CrossRef ({len(cr_needed)} S2 misses, 20 threads)")
    if cr_needed:
        cr = phase2_crossref(cr_needed, args.email, workers=20)
        for doi, abstract in cr.items():
            cache[doi] = {"abstract": abstract, "source": "CrossRef"}
        print(f"  CrossRef total: {len(cr)}  ({time.time()-t0:.1f}s)\n")
    else:
        print("  Nothing left for CrossRef\n")

    # ── Phase 3: OpenAlex for remaining misses ────────────────────────────────
    oa_needed = [d for d in to_fetch if d not in cache]
    print(f"  [Phase 3] OpenAlex batch ({len(oa_needed)} remaining)")
    if oa_needed:
        oa = phase3_openalex(oa_needed, batch_size=50)
        for doi, abstract in oa.items():
            cache[doi] = {"abstract": abstract, "source": "OpenAlex"}
        print(f"  OpenAlex total: {len(oa)}  ({time.time()-t0:.1f}s)\n")
    else:
        print("  Nothing left for OpenAlex\n")

    # ── Save cache ────────────────────────────────────────────────────────────
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    total_cached = sum(1 for v in cache.values() if v.get("abstract"))
    print(f"  Cache saved: {total_cached} abstracts ({time.time()-t0:.1f}s total)\n")

    # ── Apply enrichment ──────────────────────────────────────────────────────
    print("  [Phase 4] Applying enrichment")
    enriched = apply_enrichment(records, cache)

    # ── Save Excel ────────────────────────────────────────────────────────────
    print(f"\n  Writing enriched Excel…")
    save_enriched_xlsx(enriched, Path(args.out), Path(args.input))
    print(f"  Saved: {args.out}")

    # ── Report ────────────────────────────────────────────────────────────────
    from collections import Counter
    src_counts = Counter(r.get("Abstract_Source","") for r in enriched)
    upgraded   = sum(v for k,v in src_counts.items()
                     if k in ("SemanticScholar","CrossRef","OpenAlex"))
    lines = [
        "Abstract Enrichment Report",
        "=" * 60,
        f"Total records   : {len(enriched)}",
        f"Abstracts upgraded: {upgraded}  ({100*upgraded//len(enriched)}%)",
        "",
        "By source:",
    ]
    for src, cnt in src_counts.most_common():
        lines.append(f"  {src:<35} {cnt:>5}")
    lines += ["", f"Time elapsed: {time.time()-t0:.1f}s",
              "", "Next step:",
              "  python3 screen.py --keywords --input r2_Master_List_enriched.xlsx"]
    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")

    print(f"\n{'─'*70}")
    print(f"  Abstracts upgraded: {upgraded} / {len(enriched)}")
    print(f"  Time elapsed      : {time.time()-t0:.1f}s")
    print(f"\n  NEXT STEP:")
    print(f"    python3 screen.py --keywords --input r2_Master_List_enriched.xlsx")
    print(f"{'═'*70}\n")


if __name__ == "__main__":
    main()
