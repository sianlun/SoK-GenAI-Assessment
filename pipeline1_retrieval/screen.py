#!/usr/bin/env python3
"""
screen.py — AI-assisted title/abstract screening for the SoK review
═══════════════════════════════════════════════════════════════════════════════

WHAT THIS SCRIPT DOES
─────────────────────
This is Step 1 of the SoK pipeline. It reads the 2,162 de-duplicated records
from r2_Master_List.xlsx and classifies each one as:

    INCLUDE   → meets all inclusion criteria; goes to full-text retrieval
    EXCLUDE   → fails at least one exclusion criterion; documented with reason
    UNCERTAIN → abstract is too vague to decide; flagged for manual review

Two operating modes are available:

  --api  (default, recommended)
      Sends each record's title + abstract + keywords to Claude Haiku via the
      Anthropic API. Haiku is cheap (~$0.50 for the full 2,162-record run) and
      much more accurate than keyword matching alone on ambiguous abstracts.
      Requires ANTHROPIC_API_KEY to be set in your environment.

  --keywords
      Keyword-only mode. No API key needed. Uses a weighted scoring approach
      against the inclusion/exclusion keyword dictionaries. Faster but will
      produce more false positives. Good for a quick first pass or when you
      want a fully reproducible, deterministic run.

INCLUSION CRITERIA (ALL must be met for INCLUDE)
─────────────────────────────────────────────────
  IC1  Uses a generative AI tool: ChatGPT, GPT-4, GPT-3.5, LLM, Gemini,
       GitHub Copilot, Claude, Llama, Bard, or similar foundation model.
       (Traditional ML — SVM, ANN, CNN — without LLM does NOT qualify.)

  IC2  Assessment-related focus: automated grading, academic integrity,
       plagiarism/cheating detection, AI-generated feedback, exam generation,
       rubric evaluation, or performance scoring.

  IC3  Education context: engineering education, computer science education,
       STEM education, computing education, or technology education.
       Higher education preferred; K-12 STEM acceptable.

  IC4  Written in English with a substantive abstract (≥ 50 words).

EXCLUSION CODES
───────────────
  EC-AI      No generative AI — uses only traditional ML/rule-based systems
  EC-ASSESS  No assessment angle — pure teaching aid, tutoring, or content
             generation with no grading/integrity/feedback evaluation
  EC-EDU     Wrong education domain — medical, business, law, arts, humanities,
             language learning, corporate/workplace training
  EC-SCOPE   K-12 non-STEM, MOOCs without assessment focus, other out-of-scope
  EC-LANG    Non-English publication
  EC-TYPE    Wrong document type — editorial, letter, poster, abstract-only,
             short 1–2 page position paper without empirical/conceptual content
  EC-NOABS   No meaningful abstract (< 50 words or missing entirely)

OUTPUT FILES
────────────
  _screening_checkpoint.json   Checkpoint file — allows --resume if interrupted
  screening_results.json       Full results: every record with decision + reason
  screening_results.xlsx       Sortable Excel — same data, better for manual QC
  corpus_pending.json          INCLUDE + UNCERTAIN records only → input for
                               corpus_builder.py after your manual review

USAGE
─────
  # First run (API mode, recommended):
  export ANTHROPIC_API_KEY="sk-ant-..."
  python3 screen.py

  # Resume after interruption:
  python3 screen.py --resume

  # Keyword-only mode (no API key needed):
  python3 screen.py --keywords

  # Custom input file:
  python3 screen.py --input path/to/other.xlsx

ESTIMATED TIME & COST (API mode)
─────────────────────────────────
  Records: 2,162  |  ~300 tokens in + ~80 tokens out each
  API calls: 2,162 (one per record, sequential with 0.1 s sleep)
  Wall time: ~6–8 minutes  |  Cost: ~$0.40–0.55 (Claude Haiku)
═══════════════════════════════════════════════════════════════════════════════
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── paths ────────────────────────────────────────────────────────────────────
HERE        = Path(__file__).parent
MASTER_XLSX = HERE / "r2_Master_List.xlsx"
CHECKPOINT  = HERE / "_screening_checkpoint.json"
RESULTS_JSON= HERE / "screening_results.json"
RESULTS_XLSX= HERE / "screening_results.xlsx"
PENDING_JSON= HERE / "corpus_pending.json"

# ─── Anthropic model ──────────────────────────────────────────────────────────
MODEL = "claude-haiku-4-5-20251001"   # cheapest, fastest; sufficient for screening

# ─── inclusion / exclusion keyword dictionaries (used in keyword mode
#     and as a pre-filter in API mode to skip obvious excludes cheaply) ────────

# Generative AI signals (IC1)
GEN_AI_INCLUDE = [
    r'\bChatGPT\b',
    r'\bGPT-?4\b',
    r'\bGPT-?3\.5\b',
    r'\bGPT\b',
    r'\bLLM\b',
    r'\blarge\s+language\s+model',
    r'\bgenerative\s+AI\b',
    r'\bGenAI\b',
    r'\bGemini\b',
    r'\bGitHub\s+Copilot\b',
    r'\bCopilot\b',
    r'\bClaude\s+(?:AI|[23]|Opus|Sonnet|Haiku)\b',
    r'\bLLaMA?\b',
    r'\bLlama\s*\d',
    r'\bBard\b',
    r'\bCodex\b',
    r'\bPaLM\b',
    r'\bMistral\b',
    r'\bfoundation\s+model',
    r'\bpre-?trained\s+(?:language\s+)?model',
    r'\btransformer[-\s]based\s+(?:language\s+)?model',
]

# Assessment signals (IC2)
ASSESSMENT_INCLUDE = [
    r'\bauto\w*\s+grad\w+',
    r'\bauto\w*\s+scor\w+',
    r'\bacademic\s+integrit\w+',
    r'\bplagiar\w+',
    r'\bcheating\b',
    r'\bAI.detect\w+',
    r'\bformative\s+assess\w+',
    r'\bsummative\s+assess\w+',
    r'\bfeedback\s+generat\w+',
    r'\bAI.?feedback\b',
    r'\bautomated\s+feedback\b',
    r'\bexam\s+generat\w+',
    r'\bquiz\s+generat\w+',
    r'\brubric\b',
    r'\bgrading\b',
    r'\bpeer\s+assess\w+',
    r'\bself.?assess\w+',
    r'\bassignment\s+(?:submiss|evaluat|generat)',
    r'\bstudent\s+(?:work|submission)\s+evaluat',
    r'\bacademic\s+(?:dishonest|misconduct)',
    r'\bdetect\w+\s+(?:AI.generated|ChatGPT)',
]

# Education context signals (IC3)
EDUCATION_INCLUDE = [
    r'\bengineering\s+(?:education|student|course|program|curriculum)',
    r'\bcomputer\s+science\s+(?:education|course|student)',
    r'\bCS\s+(?:education|course|student|1\b|2\b)',
    r'\bSTEM\s+(?:education|student|course)',
    r'\bcomputing\s+education\b',
    r'\btechnology\s+education\b',
    r'\bprogramming\s+(?:course|education|class)',
    r'\bsoftware\s+engineering\s+(?:education|course)',
    r'\buniversity\b',
    r'\bundergraduate\b',
    r'\bhigher\s+education\b',
    r'\bcollege\s+(?:student|course|class)',
]

# Hard-exclude signals — if any of these hit strongly with NO generative AI,
# lean toward exclusion (used in keyword mode)
HARD_EXCLUDE_NO_AI = [
    r'\bSVM\b(?!.*(?:ChatGPT|LLM|GPT))',
    r'\bdeep\s+learning\b(?!.*(?:ChatGPT|LLM|GPT|generative))',
    r'\bneural\s+network\b(?!.*(?:ChatGPT|LLM|GPT|generative))',
    r'\bmachine\s+learning\b(?!.*(?:ChatGPT|LLM|GPT|generative|language\s+model))',
]

WRONG_DOMAIN = [
    r'\bmedical\s+(?:education|school|student)',
    r'\bnursing\s+(?:education|student)',
    r'\blaw\s+(?:school|education)',
    r'\bbusiness\s+(?:school|education)',
    r'\blanguage\s+learning\b(?!.*(?:programming|code))',
    r'\bEFL\b|\bESL\b|\bELT\b',
    r'\bcorporate\s+training\b',
    r'\bworkplace\s+training\b',
    r'\bprimary\s+school\b',
    r'\belementary\s+school\b',
    r'\bmiddle\s+school\b',
    r'\bK-?12\b(?!.*STEM)',
]


def score_record_keywords(title: str, abstract: str, keywords: str) -> dict:
    """
    Keyword-mode classifier. Returns a decision dict identical in structure
    to what the API mode returns, so the rest of the pipeline is mode-agnostic.

    Scoring:
      AI score:      count of GEN_AI_INCLUDE hits (need ≥ 1)
      Assess score:  count of ASSESSMENT_INCLUDE hits (need ≥ 1)
      Edu score:     count of EDUCATION_INCLUDE hits (need ≥ 1)
      Exclude score: count of WRONG_DOMAIN hits (any → lean exclude)
    """
    text = f"{title} {abstract} {keywords}".lower()

    ai_hits    = sum(1 for p in GEN_AI_INCLUDE    if re.search(p, text, re.I))
    assess_hits= sum(1 for p in ASSESSMENT_INCLUDE if re.search(p, text, re.I))
    edu_hits   = sum(1 for p in EDUCATION_INCLUDE  if re.search(p, text, re.I))
    excl_hits  = sum(1 for p in WRONG_DOMAIN       if re.search(p, text, re.I))

    # No abstract at all
    if len(abstract.strip()) < 50:
        return {"decision": "EXCLUDE", "confidence": 0.95,
                "code": "EC-NOABS", "reason": "Abstract missing or too short to assess."}

    # Wrong domain
    if excl_hits >= 2 and ai_hits == 0:
        return {"decision": "EXCLUDE", "confidence": 0.85,
                "code": "EC-EDU", "reason": f"Domain exclusion signals ({excl_hits}) with no GenAI signals."}

    # All three criteria hit
    if ai_hits >= 1 and assess_hits >= 1 and edu_hits >= 1:
        conf = min(0.95, 0.70 + 0.05 * (ai_hits + assess_hits + edu_hits))
        return {"decision": "INCLUDE", "confidence": round(conf, 2),
                "code": "IC1+IC2+IC3", "reason": f"GenAI({ai_hits}), Assessment({assess_hits}), Education({edu_hits}) signals all present."}

    # Missing GenAI
    if ai_hits == 0:
        return {"decision": "EXCLUDE", "confidence": 0.80,
                "code": "EC-AI", "reason": "No generative AI signals found in title/abstract/keywords."}

    # Missing assessment
    if assess_hits == 0:
        return {"decision": "EXCLUDE", "confidence": 0.75,
                "code": "EC-ASSESS", "reason": "No assessment-related signals; may be teaching/tutoring only."}

    # Missing education context
    if edu_hits == 0:
        return {"decision": "UNCERTAIN", "confidence": 0.50,
                "code": "UNCERTAIN", "reason": "GenAI and assessment signals present but education context unclear."}

    # Partial match → uncertain
    return {"decision": "UNCERTAIN", "confidence": 0.55,
            "code": "UNCERTAIN", "reason": f"Partial signals: AI={ai_hits}, Assess={assess_hits}, Edu={edu_hits}."}


def build_api_prompt(record: dict) -> str:
    """Build the screening prompt sent to Claude Haiku."""
    abstract = str(record.get("Abstract", "") or "").strip()
    abstract_excerpt = abstract[:800] if len(abstract) > 800 else abstract

    return f"""You are a systematic review screener for a Systematization of Knowledge (SoK) paper titled:
"Generative AI in Assessment within Engineering, STEM, and Computer Science Education"

INCLUSION CRITERIA — ALL must be met to classify as INCLUDE:
  IC1  Uses a generative AI tool (ChatGPT, GPT-4, GPT-3.5, LLM, Gemini, GitHub Copilot,
       Claude, Llama, Bard, Codex, PaLM, Mistral, or similar large language / foundation model).
       Traditional ML (SVM, CNN, ANN, random forest) WITHOUT an LLM does NOT qualify.
  IC2  Assessment-related focus: automated grading, academic integrity, plagiarism/cheating
       detection, AI-generated feedback on student work, exam/quiz generation, rubric
       evaluation, or performance scoring.
  IC3  Education context: engineering, computer science, STEM, computing, or technology
       education. Higher education (university/college) preferred; K-12 STEM acceptable.
  IC4  English language, substantive abstract (≥ 50 words).

EXCLUSION CODES — use the FIRST that applies:
  EC-AI      No generative AI (only traditional ML/rule-based)
  EC-ASSESS  No assessment angle (tutoring, content generation, or teaching only)
  EC-EDU     Wrong domain (medical, business, law, arts, language learning, corporate)
  EC-SCOPE   Out of scope (K-12 non-STEM, MOOCs without assessment, etc.)
  EC-LANG    Non-English
  EC-TYPE    Wrong doc type (editorial, letter, poster, <2-page abstract only)
  EC-NOABS   No meaningful abstract

RECORD TO SCREEN:
  Title:    {str(record.get("Title", ""))[:300]}
  Year:     {record.get("Year", "")}
  Source:   {str(record.get("Source_Title", ""))[:200]}
  Keywords: {str(record.get("Keywords", ""))[:300]}
  Abstract: {abstract_excerpt}

Respond with ONLY valid JSON — no markdown fences, no explanation outside the JSON:
{{"decision": "INCLUDE" or "EXCLUDE" or "UNCERTAIN", "confidence": 0.0-1.0, "code": "IC1+IC2+IC3" or one exclusion code, "reason": "one concise sentence"}}"""


def call_api(client, prompt: str, retries: int = 3) -> dict:
    """Call Claude Haiku and parse the JSON response. Retries on rate-limit errors."""
    import anthropic

    for attempt in range(retries):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=150,
                temperature=0.0,   # deterministic — important for reproducibility
                messages=[{"role": "user", "content": prompt}],
            )
            raw = response.content[0].text.strip()
            # Strip accidental markdown fences
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            result = json.loads(raw)
            # Validate required keys
            for key in ("decision", "confidence", "code", "reason"):
                if key not in result:
                    raise ValueError(f"Missing key: {key}")
            if result["decision"] not in ("INCLUDE", "EXCLUDE", "UNCERTAIN"):
                raise ValueError(f"Bad decision: {result['decision']}")
            return result

        except anthropic.RateLimitError:
            wait = 60 * (attempt + 1)
            print(f"    [Rate limit] Waiting {wait}s before retry…", flush=True)
            time.sleep(wait)

        except (json.JSONDecodeError, ValueError, KeyError) as e:
            if attempt < retries - 1:
                time.sleep(2)
                continue
            # Final fallback: mark as UNCERTAIN so we don't silently drop the record
            return {"decision": "UNCERTAIN", "confidence": 0.0,
                    "code": "PARSE-ERROR",
                    "reason": f"API response could not be parsed after {retries} attempts: {e}"}

    return {"decision": "UNCERTAIN", "confidence": 0.0,
            "code": "API-FAIL", "reason": "API call failed after all retries."}


def save_results_xlsx(results: list[dict], path: Path):
    """Write screening results to a formatted Excel workbook."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    FILLS = {
        "INCLUDE":   PatternFill("solid", fgColor="E2EFDA"),   # green
        "EXCLUDE":   PatternFill("solid", fgColor="FCE4D6"),   # orange
        "UNCERTAIN": PatternFill("solid", fgColor="FFF2CC"),   # yellow
    }
    HDR_FILL = PatternFill("solid", fgColor="1F3864")

    cols = ["Record_ID", "Decision", "Confidence", "Code", "Reason",
            "DB_Source", "Year", "Entry_Type", "Title", "Authors",
            "Source_Title", "DOI", "Keywords", "Abstract"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    # Header
    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for r_idx, rec in enumerate(results, 2):
        decision = rec.get("decision", "UNCERTAIN")
        fill = FILLS.get(decision, FILLS["UNCERTAIN"])
        row_data = [
            rec.get("Record_ID", ""),
            decision,
            rec.get("confidence", ""),
            rec.get("code", ""),
            rec.get("reason", ""),
            rec.get("DB_Source", ""),
            rec.get("Year", ""),
            rec.get("Entry_Type", ""),
            rec.get("Title", ""),
            rec.get("Authors", ""),
            rec.get("Source_Title", ""),
            rec.get("DOI", ""),
            rec.get("Keywords", ""),
            rec.get("Abstract", "")[:400] if rec.get("Abstract") else "",
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        ws.row_dimensions[r_idx].height = 18

    # Column widths
    widths = [9, 10, 10, 14, 45, 9, 6, 12, 55, 30, 35, 28, 30, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    decisions = [r.get("decision", "UNCERTAIN") for r in results]
    codes = [r.get("code", "") for r in results]

    from collections import Counter
    decision_counts = Counter(decisions)
    code_counts = Counter(codes)

    ws2["A1"] = "Screening Summary"
    ws2["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws2["A2"] = f"Total records screened: {len(results)}"
    ws2["A3"] = f"INCLUDE:   {decision_counts.get('INCLUDE', 0)}"
    ws2["A4"] = f"EXCLUDE:   {decision_counts.get('EXCLUDE', 0)}"
    ws2["A5"] = f"UNCERTAIN: {decision_counts.get('UNCERTAIN', 0)}"
    ws2["A7"] = "Exclusion Code Breakdown"
    ws2["A7"].font = Font(bold=True)
    for i, (code, cnt) in enumerate(sorted(code_counts.items()), 8):
        ws2[f"A{i}"] = code
        ws2[f"B{i}"] = cnt
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="AI-assisted abstract screening for SoK review")
    parser.add_argument("--input",    default=str(MASTER_XLSX), help="Path to master Excel file")
    parser.add_argument("--keywords", action="store_true",      help="Use keyword-only mode (no API)")
    parser.add_argument("--resume",   action="store_true",      help="Resume from checkpoint")
    parser.add_argument("--limit",    type=int, default=0,      help="Process only first N records (for testing)")
    args = parser.parse_args()

    # ── load records ──────────────────────────────────────────────────────────
    print(f"\n{'═'*70}")
    print("  SoK Screening — screen.py")
    print(f"{'═'*70}")
    print(f"  Input : {args.input}")
    print(f"  Mode  : {'Keyword-only' if args.keywords else 'Claude Haiku API'}")
    print(f"  Resume: {args.resume}")

    df = pd.read_excel(args.input, sheet_name="Unique Records", dtype=str)
    df = df.fillna("")
    records = df.to_dict("records")

    if args.limit:
        records = records[:args.limit]
        print(f"  Limit : {args.limit} records")

    total = len(records)
    print(f"  Total : {total} records to screen\n")

    # ── load checkpoint ───────────────────────────────────────────────────────
    checkpoint: dict[str, dict] = {}
    if args.resume and CHECKPOINT.exists():
        with open(CHECKPOINT) as f:
            checkpoint = json.load(f)
        print(f"  Resuming from checkpoint: {len(checkpoint)} records already done.\n")

    # ── set up API client (if needed) ─────────────────────────────────────────
    client = None
    if not args.keywords:
        try:
            import anthropic
            api_key = os.environ.get("ANTHROPIC_API_KEY")
            if not api_key:
                print("ERROR: ANTHROPIC_API_KEY not set in environment.")
                print("  Set it with: export ANTHROPIC_API_KEY='sk-ant-...'")
                print("  Or run in keyword mode: python3 screen.py --keywords")
                sys.exit(1)
            client = anthropic.Anthropic(api_key=api_key)
            print("  API client initialised. Model:", MODEL)
        except ImportError:
            print("ERROR: anthropic package not installed.")
            print("  Run: pip install anthropic")
            sys.exit(1)

    # ── main screening loop ───────────────────────────────────────────────────
    results = []
    n_include = n_exclude = n_uncertain = 0
    start_time = time.time()

    for i, rec in enumerate(records, 1):
        rec_id = str(rec.get("Record_ID", i))

        # Skip if already in checkpoint
        if rec_id in checkpoint:
            result = checkpoint[rec_id]
            results.append({**rec, **result})
            d = result.get("decision", "?")
            if d == "INCLUDE":   n_include   += 1
            elif d == "EXCLUDE": n_exclude   += 1
            else:                n_uncertain += 1
            continue

        title    = str(rec.get("Title", ""))
        abstract = str(rec.get("Abstract", ""))
        keywords = str(rec.get("Keywords", ""))

        # ── keyword pre-filter: catch obvious no-abstract / wrong-domain BEFORE
        #    spending an API call on them ──────────────────────────────────────
        if not args.keywords:
            if len(abstract.strip()) < 50:
                result = {"decision": "EXCLUDE", "confidence": 0.95,
                          "code": "EC-NOABS", "reason": "No meaningful abstract."}
            elif any(re.search(p, f"{title} {abstract}", re.I) for p in WRONG_DOMAIN) \
                 and not any(re.search(p, f"{title} {abstract} {keywords}", re.I) for p in GEN_AI_INCLUDE):
                result = {"decision": "EXCLUDE", "confidence": 0.85,
                          "code": "EC-EDU", "reason": "Domain exclusion signals with no GenAI signals (pre-filter)."}
            else:
                # Full API call
                prompt = build_api_prompt(rec)
                result = call_api(client, prompt)
                time.sleep(0.12)   # ~8 req/s → well under 2000 RPM rate limit
        else:
            # Keyword-only
            result = score_record_keywords(title, abstract, keywords)

        # ── accumulate ────────────────────────────────────────────────────────
        checkpoint[rec_id] = result
        results.append({**rec, **result})

        d = result["decision"]
        if d == "INCLUDE":   n_include   += 1
        elif d == "EXCLUDE": n_exclude   += 1
        else:                n_uncertain += 1

        # Progress log every 50 records
        if i % 50 == 0 or i == total:
            elapsed = time.time() - start_time
            rate = i / elapsed if elapsed > 0 else 0
            eta  = (total - i) / rate if rate > 0 else 0
            print(f"  [{i:>4}/{total}] IN={n_include}  EX={n_exclude}  UN={n_uncertain}"
                  f"  |  {rate:.1f} rec/s  ETA {eta/60:.1f} min", flush=True)

        # Save checkpoint every 50 records
        if i % 50 == 0:
            with open(CHECKPOINT, "w") as f:
                json.dump(checkpoint, f)

    # ── final checkpoint save ─────────────────────────────────────────────────
    with open(CHECKPOINT, "w") as f:
        json.dump(checkpoint, f)

    # ── write outputs ─────────────────────────────────────────────────────────
    print(f"\n{'─'*70}")
    print(f"  Screening complete.")
    print(f"    INCLUDE  : {n_include:>5}  ({100*n_include/total:.1f}%)")
    print(f"    EXCLUDE  : {n_exclude:>5}  ({100*n_exclude/total:.1f}%)")
    print(f"    UNCERTAIN: {n_uncertain:>5}  ({100*n_uncertain/total:.1f}%)")
    print(f"    Total    : {total}")

    # Full results JSON
    with open(RESULTS_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\n  Saved: {RESULTS_JSON}")

    # Formatted Excel
    save_results_xlsx(results, RESULTS_XLSX)
    print(f"  Saved: {RESULTS_XLSX}")

    # Pending corpus — INCLUDE + UNCERTAIN → send to corpus_builder.py
    pending = [r for r in results if r.get("decision") in ("INCLUDE", "UNCERTAIN")]
    pending_out = [
        {
            "record_id":   r.get("Record_ID", ""),
            "decision":    r.get("decision", ""),
            "confidence":  r.get("confidence", ""),
            "code":        r.get("code", ""),
            "reason":      r.get("reason", ""),
            "title":       r.get("Title", ""),
            "authors":     r.get("Authors", ""),
            "year":        r.get("Year", ""),
            "doi":         r.get("DOI", ""),
            "source":      r.get("Source_Title", ""),
            "db_source":   r.get("DB_Source", ""),
            "abstract":    r.get("Abstract", ""),
            "keywords":    r.get("Keywords", ""),
            # Fields for manual review — to be filled in corpus_builder.py
            "manual_decision":  "",    # INCLUDE / EXCLUDE after your review
            "manual_note":      "",
        }
        for r in pending
    ]
    with open(PENDING_JSON, "w", encoding="utf-8") as f:
        json.dump(pending_out, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {PENDING_JSON}  ({len(pending_out)} records for manual review)")

    print(f"\n  NEXT STEP: Review screening_results.xlsx, especially UNCERTAIN records.")
    print(f"  Then run: python3 corpus_builder.py")
    print(f"{'═'*70}\n")


if __name__ == "__main__":
    main()
